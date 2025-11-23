# -*- coding: utf-8 -*-
"""
Created on Mon Jun 23 16:19:05 2025

@author: djmiller
"""



import numpy as np
import xlwings as xw
import pandas as pd
import multiprocessing as mp
import itertools
import copy

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows  # Correct import

import os

from anchor_pro.equipment import EquipmentModel, SMSHardwareAttachment
from anchor_pro.report import EquipmentReport

from dataclasses import dataclass
from typing import Literal, Optional

pd.set_option('future.no_silent_downcasting', True)


class ExcelTablesImporter:
    NA_VALUES = ['NA', 'N/A', '-', 'None', ""]

    def __init__(self, excel_path):
        print('Reading data from Excel')
        self.path = excel_path
        wb = xw.Book(self.path)

        # Import Project Info from Named Cells on Project Worksheet
        project_sheet = wb.sheets['Project']
        self.project_info = {}
        for name in wb.names:
            if name.name.startswith('_xlfn') or name.name.startswith('_xlpm') or name.name.startswith('_xleta'):
                continue
            try:
                if name.refers_to_range.sheet == project_sheet:
                    value = name.refers_to_range.value
                    self.project_info[name.name] = value
            except Exception as e:
                print(f"Skipping {name.name} due to error: {e}")

        # Import Excel Tables
        self.df_equipment = None
        self.df_base_geometry = None
        self.df_wall_geometry = None
        self.df_concrete = None
        self.df_walls = None
        self.df_anchors = None
        self.df_brackets_catalog = None
        self.bracket_group_list = None
        self.df_product_groups = None
        self.df_wood = None

        # Table References (sheet name, table NW cell name, instance attribute name)
        table_references = [('Equipment', 'tbl_equipment', 'df_equipment'),
                            ('Base Geometry', 'tbl_base_geometry', 'df_base_geometry'),
                            ('Wall Geometry', 'tblBrackets', 'df_wall_geometry'),
                            ('Fastener Patterns', 'tblBacking', 'df_fasteners'),
                            ('Concrete', 'tbl_concrete', 'df_concrete'),
                            ('Wood','tbl_wood', 'df_wood'),
                            ('Walls', 'tblWalls', 'df_walls'),
                            ('Anchors', 'tbl_anchors', 'df_anchors'),
                            ('Wood Fasteners', 'tbl_wood_fasteners', 'df_wood_fasteners'),
                            ('SMS', 'tblSMS', 'df_sms'),
                            ('Brackets', 'tblBracketCatalog', 'df_brackets_catalog'),
                            ('Anchor Product Groups', 'tblProductGroups', 'df_product_groups')]

        for (sheet_name, tbl_name, df_name) in table_references:
            sheet = wb.sheets[sheet_name]
            table_cell = sheet.range(tbl_name)
            start_address = table_cell.get_address(0, 0, include_sheetname=True, external=False)
            df = sheet.range(start_address).expand().options(pd.DataFrame,
                                                             header=1,
                                                             index=False,
                                                             expand='table').value
            setattr(self, df_name, df)

        # Replace undesired inputs
        self.df_equipment = self.df_equipment.replace(ExcelTablesImporter.NA_VALUES, None)

        # Adjust Indicies on Certain Tables
        for df in [self.df_walls, self.df_product_groups, self.df_fasteners]:
            df.set_index(df.columns[0], inplace=True)

        # Import Bracket Product Group Names
        self.bracket_group_list = wb.names['bracket_groups_list'].refers_to_range.value


class ProjectController:
    TABLE_COLUMNS = {
        # GENERAL INFO
        'Item': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Group': {'units': None, 'width': 12, 'alignment': 'l', 'style': None},
        'Wp': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Fp': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        # BASE CONCRETE ANCHOR RESULTS
        'Base Anchor': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Base Anchor Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Base Thickness Check': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Base Spacing and Edge Checks': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Base Steel Tensile Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Concrete Tension Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Anchor Pullout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Side Face Blowout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Bond Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Steel Shear Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Shear Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Shear Pryout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Tension-Shear Interaction': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Anchor OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Base Anchor': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # HARDWARE FASTENERS
        'Hardware SMS': {'units': None, 'width': None, 'alignment': 'l', 'style': None},
        'Hardware SMS Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Hardware SMS Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Hardware SMS Tension DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Hardware SMS Shear DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Hardware SMS OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Hardware SMS': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # BASE STRAPS
        'Base Strap': {'units': None, 'width': 16, 'alignment': 'l', 'style': None},
        'Maximum Base Strap Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Base Strap DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Strap OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL BRACKETS
        'Wall Bracket': {'units': None, 'width': 16, 'alignment': 'l', 'style': None},
        'Maximum Bracket Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Bracket DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Bracket OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL SMS ANCHORS
        'Wall SMS': {'units': None, 'width': 8, 'alignment': 'l', 'style': None},
        'Wall SMS Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall SMS Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall SMS Tension DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS Shear DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Wall SMS': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # WALL WOOD FASTENER
        'Wall Fastener': {'units': None, 'width': 8, 'alignment': 'l', 'style': None},
        'Wall Fastener Max Force': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        #'Wall Fastner Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall Fastener DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Fastener OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL CONCRETE ANCHOR
        'Wall Anchor': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Wall Anchor Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall Thickness Check': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Wall Spacing and Edge Checks': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Wall Steel Tensile Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Concrete Tension Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Anchor Pullout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Side Face Blowout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Bond Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Steel Shear Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Shear Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Shear Pryout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Tension-Shear Interaction': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Anchor OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Wall Anchor': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'}
    }

    def __init__(self, excel_path, output_dir):
        # Inputs
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.excel_tables = None
        self.df_equipment = None  # Merge of equipment, base geometry, wall geometry, concrete tables
        self.model_inputs_dictionaries = {}
        self.get_model_inputs_from_excel_tables()

        # Multiprocessing pool
        self.pool = mp.Pool(processes=mp.cpu_count()) if self.excel_tables.project_info[
            'use_parallel_processing'] else None

        # Outputs
        self.items_for_report = {}
        self.governing_items = None
        self.group_dict = None
        self.df_results = None

    def input_validation(self):
        """ Function to validate all required user inputs are provided before attempting analysis"""
        pass

    @staticmethod
    def append_results(results_lists, model, is_optimum_base=False, is_optimum_wall=False):
        # Create blank row
        for key, results in results_lists.items():
            results.append('')

        # Populate item data
        results_lists['Item'][-1] = model.equipment_id
        results_lists['Group'][-1] = model.group
        results_lists['Wp'][-1] = model.Wp
        results_lists['Fp'][-1] = model.Fp

        # Populate base anchor data
        if model.base_anchors is not None:
            results_lists['Base Anchor'][-1] = model.base_anchors.anchor_id
            results_lists['Base Thickness Check'][-1] = \
                "OK" if model.base_anchors.spacing_requirements['slab_thickness_ok'] else "NG"
            results_lists['Base Spacing and Edge Checks'][-1] = \
                "OK" if model.base_anchors.spacing_requirements['edge_and_spacing_ok'] else "NG"

            for limit_state in ['Steel Tensile Strength',
                                'Concrete Tension Breakout',
                                'Anchor Pullout',
                                'Side Face Blowout',
                                'Bond Strength',
                                'Steel Shear Strength',
                                'Shear Pryout']:

                if model.omit_analysis:
                    result = 'NA'
                elif limit_state not in model.base_anchors.results.index:
                    result = 'NA'
                else:
                    result = model.base_anchors.results.loc[limit_state, 'Utilization']
                results_lists['Base ' + limit_state][-1] = result

            if model.omit_analysis:
                results_lists['Base Anchor Max Tension'][-1] = 'NA'
                results_lists['Base Shear Breakout'][-1] = 'NA'
            else:
                results_lists['Base Anchor Max Tension'][-1] = model.base_anchors.anchor_force_results[:, 0].max()
                cases = list(
                    set(model.base_anchors.shear_breakout_long_name_to_short_name_map.keys()) & set(
                        model.base_anchors.results.index))
                max_breakout = model.base_anchors.results.loc[cases, 'Utilization'].max()
                results_lists['Base Shear Breakout'][-1] = max_breakout

            results_lists['Base Tension-Shear Interaction'][
                -1] = 'NA' if model.base_anchors.DCR is None else model.base_anchors.DCR
            results_lists['Base Anchor OK'][
                -1] = False if model.base_anchors.DCR is None else (model.base_anchors.DCR < 1)

            results_lists['Optimum Base Anchor'][-1] = is_optimum_base

        # Populate Hardware Bracket Data  (Based on Governing Base Plate Connection or Wall Brace Connection
        floor_plate_sms_list = [plate.connection.anchors_obj for plate in model.floor_plates
                                if isinstance(plate.connection, SMSHardwareAttachment)
                                and plate.connection.anchors_obj.results]
        if floor_plate_sms_list:
            model.update_element_resultants(model.governing_solutions['base_anchor_tension']['sol'])
            floor_plate_sms = max(floor_plate_sms_list, key=lambda x: x.max_dcr())
            floor_plate_dcr = floor_plate_sms.max_dcr()
        else:
            floor_plate_sms = None
            floor_plate_dcr = -np.inf

        bracket_sms_list = [bracket.connection.anchors_obj for bracket in model.wall_brackets
                            if isinstance(bracket.connection, SMSHardwareAttachment)
                            and bracket.connection.anchors_obj.results]
        if bracket_sms_list:
            model.update_element_resultants(model.governing_solutions['wall_bracket_tension']['sol'])
            bracket_sms = max(bracket_sms_list, key=lambda x: x.max_dcr())
            bracket_sms_dcr = bracket_sms.max_dcr()
        else:
            bracket_sms = None
            bracket_sms_dcr = -np.inf

        if floor_plate_sms or bracket_sms:
            if floor_plate_dcr > bracket_sms_dcr:
                model.update_element_resultants(model.governing_solutions['base_anchor_tension']['sol'])
                governing_sms = floor_plate_sms
            else:
                governing_sms = bracket_sms

            results_lists['Hardware SMS'][-1] = governing_sms.screw_size
            results_lists['Hardware SMS Max Tension'][-1] = governing_sms.results['Tension Demand']
            results_lists['Hardware SMS Max Shear'][-1] = governing_sms.results['Shear Demand']
            results_lists['Hardware SMS Tension DCR'][-1] = governing_sms.results['Tension DCR']
            results_lists['Hardware SMS Shear DCR'][-1] = max(governing_sms.results['Shear X DCR'],
                                                              governing_sms.results['Shear Y DCR'])
            results_lists['Hardware SMS OK'][-1] = bool(governing_sms.results['OK'])

        # Populate Base Straps Data
        if model.base_straps and model.omit_analysis:
            results_lists['Base Strap'][-1] = model.base_straps[0].bracket_id
            results_lists['Maximum Base Strap Tension'][-1] = 'NA'
            results_lists['Base Strap DCR'][-1] = 'NA'
            results_lists['Base Strap OK'][-1] = 'NA'
        elif model.base_straps:
            governing_strap = max(model.base_straps, key=lambda x: x.tension_dcr)
            results_lists['Base Strap'][-1] = governing_strap.bracket_id
            results_lists['Maximum Base Strap Tension'][-1] = governing_strap.brace_force
            results_lists['Base Strap DCR'][-1] = governing_strap.tension_dcr
            results_lists['Base Strap OK'][-1] = bool(governing_strap.tension_dcr <= 1) if isinstance(governing_strap.tension_dcr, (int, float)) else governing_strap.tension_dcr

        # Populate Wall Bracket Data
        if model.wall_brackets:
            results_lists['Wall Bracket'][-1] = model.wall_brackets[0].bracket_id
            if not model.omit_analysis:
                idx_bracket, idx_theta = np.unravel_index(np.argmax(model.wall_bracket_forces[:, :, 0]),
                                                          model.wall_bracket_forces[:, :, 0].shape)
                asd_lrfd_conversion = model.asd_lrfd_ratio if model.wall_brackets[0].capacity_method == 'ASD' else 1
                results_lists['Maximum Bracket Tension'][-1] = np.max(
                    model.wall_bracket_forces[idx_bracket, idx_theta, 0])*asd_lrfd_conversion
                dcr = model.wall_brackets[idx_bracket].tension_dcr
                results_lists['Bracket DCR'][-1] = dcr
                results_lists['Bracket OK'][-1] = bool(dcr <= 1) if isinstance(dcr, (int, float)) else True

        # Populate Wall Anchor Data
        if model.wall_type == 'Metal Stud':
            # Determine wall_anchors object with maximum anchor tension
            walls_with_anchors = [(b.anchors_obj, b.anchors_obj.results['Tension DCR']) for
                                  b in model.wall_backing if
                                  b.anchors_obj is not None and b.anchors_obj.results]
            wall_anchors = max(walls_with_anchors, key=lambda x: x[1], default=(None, -np.inf))[0]

            if wall_anchors:
                results_lists['Wall SMS'][-1] = wall_anchors.screw_size
                if not model.omit_analysis:
                    results_lists['Wall SMS Max Tension'][-1] = wall_anchors.results['Tension Demand']
                    results_lists['Wall SMS Max Shear'][-1] = wall_anchors.results['Shear Demand']
                    results_lists['Wall SMS Tension DCR'][-1] = wall_anchors.results['Tension DCR']
                    results_lists['Wall SMS Shear DCR'][-1] = wall_anchors.results['Shear DCR']
                    results_lists['Wall SMS DCR'][-1] = wall_anchors.DCR
                    results_lists['Wall SMS OK'][-1] = wall_anchors.results['OK']

                results_lists['Optimum Wall SMS'][-1] = is_optimum_wall
        elif model.wall_type == 'Wood Stud':
            # Determine wall_anchors object with maximum anchor tension
            anchor_objects = [(b.anchors_obj, b.anchors_obj.DCR) for
                              b in model.wall_backing if
                              b.anchors_obj is not None]
            wall_anchors = max(anchor_objects, key=lambda x: x[1], default=(None, -np.inf))[0]

            if wall_anchors:
                results_lists['Wall Fastener'][-1] = wall_anchors.fastener_id
                results_lists['Wall Fastener Max Force'][-1] = wall_anchors.Tu_max
                results_lists['Wall Fastener DCR'][-1] = wall_anchors.DCR
                results_lists['Wall Fastener OK'][-1] = wall_anchors.DCR < 1

        elif model.wall_type == 'Concrete':
            # Determine wall_anchors object with maximum anchor tension
            walls_with_anchors = [(wall, wall.DCR) for wall in model.wall_anchors.values() if
                                  wall is not None]
            wall_anchors = max(walls_with_anchors, key=lambda x: x[1], default=(None, -np.inf))[0]

            if wall_anchors:
                results_lists['Wall Anchor'][-1] = wall_anchors.anchor_id
                results_lists['Wall Thickness Check'][-1] = \
                    "OK" if wall_anchors.spacing_requirements['slab_thickness_ok'] else "NG"
                results_lists['Wall Spacing and Edge Checks'][-1] = \
                    "OK" if wall_anchors.spacing_requirements['edge_and_spacing_ok'] else "NG"

                for limit_state in ['Steel Tensile Strength',
                                    'Concrete Tension Breakout',
                                    'Anchor Pullout',
                                    'Side Face Blowout',
                                    'Bond Strength',
                                    'Steel Shear Strength',
                                    'Shear Pryout']:

                    if model.omit_analysis:
                        result = 'NA'
                    elif limit_state not in wall_anchors.results.index:
                        result = 'NA'
                    else:
                        result = wall_anchors.results.loc[limit_state, 'Utilization']
                    results_lists['Wall ' + limit_state][-1] = result

                if model.omit_analysis:
                    results_lists['Wall Anchor Max Tension'][-1] = 'NA'
                    results_lists['Wall Shear Breakout'][-1] = 'NA'
                else:
                    results_lists['Wall Anchor Max Tension'][-1] = wall_anchors.anchor_force_results[:,
                                                                   0].max()
                    cases = list(
                        set(wall_anchors.shear_breakout_long_name_to_short_name_map.keys()) & set(wall_anchors.results.index))
                    max_breakout = wall_anchors.results.loc[cases, 'Utilization'].max()
                    results_lists['Wall Shear Breakout'][-1] = max_breakout

                results_lists['Wall Tension-Shear Interaction'][
                    -1] = 'NA' if wall_anchors.DCR is None else wall_anchors.DCR
                results_lists['Wall Anchor OK'][
                    -1] = False if wall_anchors.DCR is None else (wall_anchors.DCR < 1)

                results_lists['Optimum Wall Anchor'][-1] = is_optimum_wall

    def get_model_inputs_from_excel_tables(self):
        """Performs appropriate lookups across dataframes (equipment table, base geometry, etc.) and assembles
        a dictionary of data series representing the required inputs for each equipment model.

        the dictionary contains the following keys:
        'equipment_data'
        'concrete_data'
        'base_geometry'
        'anchor_geometry'

        """
        # Read Excel File and Import Data Tables as DataFrames
        self.excel_tables = ExcelTablesImporter(self.excel_path)

        # Merge Tables
        df_equipment = self.excel_tables.df_equipment
        df_equipment = df_equipment[df_equipment['omit_calc'] != True]

        # Merge Base Geometry
        df_equipment = df_equipment.merge(self.excel_tables.df_base_geometry,
                                          left_on='base_geometry', right_on='Pattern Name',
                                          how='left')
        df_equipment = df_equipment.drop(columns=['Pattern Name'])

        # Merge Wall Geometry
        df_equipment = df_equipment.merge(self.excel_tables.df_wall_geometry,
                                          left_on='wall_geometry', right_on='Pattern Name',
                                          how='left', suffixes=('_base', '_wall'))
        df_equipment = df_equipment.drop(columns=['Pattern Name'])

        # Merge Base Material Data
        # todo: [New Base Materials] merge additional tables for other material types
        df_equipment = df_equipment.merge(self.excel_tables.df_concrete,
                                          left_on='base_mat_id', right_on='slab_id',
                                          how='left', suffixes=('_eq', '_base'))
        df_equipment = df_equipment.merge(self.excel_tables.df_wood,
                                          left_on='base_mat_id', right_on='wood_id',
                                          how='left', suffixes=('_eq', '_base'))

        # Merge Wall Material Data
        df_equipment = df_equipment.merge(self.excel_tables.df_walls,
                                          left_on='wall_id', right_on='wall_id',
                                          how='left', suffixes=('_base', '_wall'))
        df_equipment = df_equipment.merge(self.excel_tables.df_concrete,
                                          left_on='slab_id_wall', right_on='slab_id',
                                          how='left', suffixes=('_base', '_wall'))

        df_equipment['stud_or_blocking_key'] = (df_equipment['blocking_id'].fillna(df_equipment['stud_id']).astype(str))
        df_equipment = df_equipment.merge(
            self.excel_tables.df_wood,
            left_on='stud_or_blocking_key',right_on='wood_id',
            how='left', suffixes=('_base','_wall')).drop(columns=['stud_or_blocking_key'])

        ''' The decision is made deliberately not to merge anchor data or bracket data.
        These tables are independently quite large, and since checking by group may be selected,
        it seems preferable not to explicitly include these tables in df_equipment.'''

        self.df_equipment = df_equipment

    def run_analysis(self):

        df_results = pd.DataFrame()

        if self.excel_tables.project_info['use_parallel_processing']:
            # Parallel Processing
            print('\n')
            print('#' * 40)
            print('Beginning analysis (parallel processing)')

            items_and_results_list = self.pool.starmap(self.create_model_and_analyze_item,
                                                       [(self.excel_tables.project_info, equipment_data,
                                                         self.excel_tables.df_anchors,
                                                         self.excel_tables.df_wood_fasteners,
                                                         self.excel_tables.df_product_groups,
                                                         self.excel_tables.df_brackets_catalog,
                                                         self.excel_tables.bracket_group_list,
                                                         self.excel_tables.df_fasteners,
                                                         self.excel_tables.df_sms) for index, equipment_data in
                                                        self.df_equipment.iterrows()])

            for (model, d) in items_and_results_list:
                self.items_for_report[model.equipment_id] = model
                df_results = pd.concat([df_results, pd.DataFrame(d)], ignore_index=True)

        else:
            # Serial Processing
            print('\n')
            print('#' * 40)
            print('Beginning analysis (serial processing)')
            print('#' * 40)

            results_list = []
            for index, equipment_data in self.df_equipment.iterrows():
                model, results = self.create_model_and_analyze_item(self.excel_tables.project_info, equipment_data,
                                                                    self.excel_tables.df_anchors,
                                                                    self.excel_tables.df_wood_fasteners,
                                                                    self.excel_tables.df_product_groups,
                                                                    self.excel_tables.df_brackets_catalog,
                                                                    self.excel_tables.bracket_group_list,
                                                                    self.excel_tables.df_fasteners,
                                                                    self.excel_tables.df_sms)
                results_list.append(results)
                self.items_for_report[model.equipment_id] = model

            for d in results_list:
                df_results = pd.concat([df_results, pd.DataFrame(d)], ignore_index=True)

        self.get_governing_by_group()
        print('Analysis Complete')

        print('\n')
        print('#' * 40)
        print('Compiling results')
        print('#' * 40)

        # Create a workbook and select the active worksheet
        self.df_results = df_results
        self.create_excel_summary_table(df_results)

    @staticmethod
    def create_model_and_analyze_item(project_info, equipment_data, anchor_catalog, wood_fastener_catalog, anchor_product_groups,
                                      wall_bracket_catalog, bracket_groups_list, df_fasteners, df_sms):
        """ Function to run analysis of single model"""
        results_lists = {key: [] for key in ProjectController.TABLE_COLUMNS}

        # Create Model
        print(f'Creating {equipment_data["equipment_id"]} Model')
        model = EquipmentModel()
        model.set_model_data(project_info, equipment_data, df_fasteners, df_sms)
        model.calculate_fp()
        model.calculate_factored_loads()

        def get_anchor_product_list(group_name, member_type='Slab'):
            if group_name is None:
                return None

            anchor_products = anchor_product_groups[
                anchor_product_groups[group_name] == True].index
            filtered_anchors_df = anchor_catalog[anchor_catalog['product'].isin(anchor_products)]
            if member_type == 'Slab':
                filtered_anchors_df = filtered_anchors_df[filtered_anchors_df['slab_ok']]
            elif member_type == 'Filled Deck':
                filtered_anchors_df = filtered_anchors_df[filtered_anchors_df['deck_top_ok']]
            return filtered_anchors_df['anchor_id'].tolist()

        def get_sms_list(product_mode):
            if product_mode == 'Default':
                return ['No. 12']
            else:
                return ['No. 14', 'No. 12', 'No. 10', 'No. 8', 'No. 6']

        def get_bracket_group_list():
            idx = bracket_groups_list.index(model.bracket_group) + 1
            catalog_group = f'group_{idx}'
            filtered_bracket_df = wall_bracket_catalog[wall_bracket_catalog[catalog_group] == True]
            return filtered_bracket_df['bracket_id'].tolist()

        def get_wood_fastener_group_list():
            pass

        '''Collect Product Lists'''
        # Initialize Lists
        product_lists = {'base_anchor_list': [None],
                         'cxn_anchor_list': [None],
                         'bracket_list': [None],
                         'wall_anchor_list': [None]}

        product_applicable = {'base_anchor_list': model.base_anchors is not None,
                              'cxn_anchor_list': any([p.connection is not None for p in model.floor_plates]) or any(
                                  [b.connection is not None for b in model.wall_brackets]),
                              'bracket_list': len(model.wall_brackets) > 0,
                              'wall_anchor_list': any([v is not None for k, v in model.wall_anchors.items()]+
                                                      [b.anchors_obj is not None for b in model.wall_backing])}

        # Product List Parameters
        '''(list_object,
        material_applicable, 
        product_mode, 
        specified product, product group, 
        group_function, args, kwargs'''
        list_parameters = [
            # Base Concrete Anchors
            ('base_anchor_list',
             model.base_material == 'Concrete',  # Replace with applicable to concrete when adding other base anchor types
             project_info['base_concrete_product_mode'],
             model.base_anchor_id, model.base_anchor_group,
             get_anchor_product_list, [model.base_anchor_group], {'member_type': model.profile_base}),
            # Hardware Connection SMS
            ('cxn_anchor_list',
             True,
             project_info['hardware_sms_product_mode'],
             model.cxn_sms_id, True,
             get_sms_list, [project_info['hardware_sms_product_mode']], {}),
            # Wall Brackets
            ('bracket_list',
             True,
             project_info['wall_bracket_product_mode'],
             model.bracket_id, model.bracket_group,
             get_bracket_group_list, [], {}),
            # Wall Concrete/CMU Anchors
            ('wall_anchor_list',
             model.wall_type in ['Concrete', 'CMU'],
             project_info['wall_concrete_product_mode'],
             model.wall_anchor_id, model.wall_anchor_group,
             get_anchor_product_list, [model.wall_anchor_group], {'member_type': 'Slab'}),
            # Wall SMS Anchors
            ('wall_anchor_list',
             model.wall_type == 'Metal Stud',
             project_info['wall_sms_product_mode'],
             model.wall_sms_id, True,
             get_sms_list, [project_info['wall_sms_product_mode']], {}),
            ('wall_anchor_list',
             model.wall_type == 'Wood Stud',
             project_info['wall_sms_product_mode'],
             model.wall_fastener_id, model.wall_fastener_group,
             get_wood_fastener_group_list, [project_info['wall_sms_product_mode']], {})
        ]

        for list_name, material_applicable, mode, product, group, func, args, kwargs in list_parameters:
            if product_applicable[list_name] and material_applicable:
                if mode in ['Default', 'Specified Product'] and product:
                    product_lists[list_name] = [product]
                elif mode in ['Default', 'Product Group'] and group:
                    product_lists[list_name] = func(*args, **kwargs)

        # Verify That product is provided for all required items
        for list_name, product_list in product_lists.items():
            if product_applicable[list_name] and product_list == [None]:
                raise Exception(f"Must specify product or group for {model.equipment_id}, {list_name}")

        # Initialize Results Management Parameters
        optimum_base_cost = float('inf')
        optimum_wall_cost = float('inf')
        optimum_screw_index = 0
        optimum_base_results_index = None
        optimum_wall_results_index = None
        results_index = 0
        is_optimum_base = False
        is_optimum_wall = False
        item_for_report = None
        initial_solution_cache = None

        for base_anchor_id, bracket_id, wall_anchor_id, cxn_anchor_id in \
                itertools.product(product_lists['base_anchor_list'],
                                  product_lists['bracket_list'],
                                  product_lists['wall_anchor_list'],
                                  product_lists['cxn_anchor_list']):
            print(f'Analyzing {equipment_data["equipment_id"]} with '
                  f'base anchor: {base_anchor_id}, '
                  f'wall bracket: {bracket_id}, '
                  f'wall anchor: {wall_anchor_id}')

            base_anchor_data = None
            base_strap_data = None
            bracket_data = None
            wall_anchor_data = None

            if base_anchor_id:
                if model.base_material == 'Concrete':
                    base_anchor_data = anchor_catalog[anchor_catalog['anchor_id'] == base_anchor_id].iloc[0]
            if len(model.base_straps) > 0:
                base_strap_data = wall_bracket_catalog[wall_bracket_catalog['bracket_id'] == model.base_strap].iloc[0]
            if bracket_id:
                bracket_data = wall_bracket_catalog[wall_bracket_catalog['bracket_id'] == bracket_id].iloc[0]
            if wall_anchor_id:
                if model.wall_type in ['Concrete', 'CMU']:
                    wall_anchor_data = anchor_catalog[anchor_catalog['anchor_id'] == wall_anchor_id].iloc[0]
                elif model.wall_type == 'Metal Stud':
                    wall_anchor_data = wall_anchor_id  # For SMS, only "data" is anchor size
                elif model.wall_type == 'Wood Stud':
                    wall_anchor_data = wood_fastener_catalog[wood_fastener_catalog['fastener_id']==wall_anchor_id].iloc[0]


            model.set_product_data_and_analyze(base_anchor_data=base_anchor_data,
                                               base_strap_data=base_strap_data,
                                               bracket_data=bracket_data,
                                               wall_anchor_data=wall_anchor_data,
                                               hardware_screw_size=cxn_anchor_id,
                                               initial_solution_cache=initial_solution_cache)

            # TODO FINISH RESULTS STORE
            # results = ResultsStore()

            # Determine Optimum Base Anchor
            if model.installation_type in ['Base Anchored',
                                           'Wall Brackets'] and model.base_anchors is not None:
                current_cost = base_anchor_data['cost_rank']
                anchor_applicable = False if model.base_anchors.DCR is None else model.base_anchors.DCR < 1
                is_optimum_base = anchor_applicable and current_cost < optimum_base_cost
                if is_optimum_base:
                    # Update the results_lists to mark the previous optimum as not applicable
                    if optimum_base_results_index is not None:
                        results_lists['Optimum Base Anchor'][optimum_base_results_index] = False

                    # Update the current optimum
                    optimum_base_cost = current_cost
                    optimum_base_results_index = results_index
                    item_for_report = copy.deepcopy(model)  # Deep copy to preserve the state

            # Determine Optimum Wall Anchor
            if model.installation_type in ['Wall Brackets', 'Wall Mounted']:
                if model.wall_type == 'Metal Stud':
                    current_screw_index = product_lists['wall_anchor_list'].index(wall_anchor_id)
                    anchor_applicable = all(
                        [backing.anchors_obj.results['OK'] for backing in model.wall_backing if
                         backing.anchors_obj is not None and backing.anchors_obj.results])
                    is_optimum_wall = anchor_applicable and current_screw_index >= optimum_screw_index

                    if is_optimum_wall:
                        # Update the results_lists to mark the previous optimum as not applicable
                        if optimum_wall_results_index is not None:
                            results_lists['Optimum Wall SMS'][optimum_wall_results_index] = False

                        # Update the current optimum
                        optimum_screw_index = current_screw_index
                        optimum_wall_results_index = results_index
                        if model.base_anchors is None:
                            item_for_report = copy.deepcopy(model)  # Deep copy to preserve the state
                elif model.wall_type == 'Concrete':
                    current_cost = wall_anchor_data['cost_rank']
                    dcr_list = [anchors.DCR for anchors in model.wall_anchors.values() if anchors]
                    dcr_ok = [dcr < 1 for dcr in dcr_list if dcr]
                    anchor_applicable = all(dcr_ok) if dcr_list else False
                    is_optimum_wall = anchor_applicable and current_cost < optimum_wall_cost
                    if is_optimum_wall:
                        # Update the results_lists to mark the previous optimum as not applicable
                        if optimum_wall_results_index is not None:
                            results_lists['Optimum Wall Anchor'][optimum_wall_results_index] = False

                        # Update the current optimum
                        optimum_wall_cost = current_cost
                        optimum_results_index = results_index
                        item_for_report = copy.deepcopy(model)  # Deep copy to preserve the state
                # todo: select optimum base plate or bracket attachment sms
            if initial_solution_cache is None and not model.omit_analysis:
                initial_solution_cache = model.equilibrium_solutions
            results_index += 1
            ProjectController.append_results(results_lists, model, is_optimum_base=is_optimum_base,
                                             is_optimum_wall=is_optimum_wall)

        if item_for_report is None:
            item_for_report = copy.deepcopy(model)



        return item_for_report, results_lists

    def check_product_group(self):
        """Checks all anchor products in the specified product group and reports the "minimum" anchor"""

    def create_excel_summary_table(self, df_results):
        print('Creating summary table')
        # Export and Format Results Summary Table (using Openpyxl)
        file_path = os.path.join(self.output_dir, 'Results Summary.xlsx')

        wb = Workbook()
        ws = wb.active

        # Populate the sheet with DataFrame excel_tables
        for r in dataframe_to_rows(df_results, index=False, header=True):
            ws.append(r)

        # Predefine Fill Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FF7276", end_color="FF7276", fill_type="solid")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        def style_dcr(c):
            if isinstance(c.value, (int, float)) and c.value > 1:
                c.fill = red_fill

        def style_condition(c):
            if c.value == 'NG':
                c.fill = red_fill
            elif c.value == 'OK':
                c.fill = green_fill

        def style_result(c):
            if c.value:
                c.fill = green_fill
            else:
                c.fill = gray_fill

        def style_optimum(c):
            if c.value:
                c.fill = green_fill
            else:
                c.fill = gray_fill

        def style_header(c):
            c.fill = header_fill
            c.font = Font(color="FFFFFF")  # White Font
            c.border = thin_border
            c.alignment = Alignment(horizontal='left', vertical='bottom', text_rotation=45)

        def style_general(c):
            # Only apply number formatting if it's a number, not a boolean
            if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                c.number_format = '0.00'
            c.border = thin_border
            if c.value == 'NA' or c.value == "":
                c.fill = gray_fill

        style_functions = {'dcr': style_dcr,
                           'optimum': style_optimum,
                           'result': style_result,
                           'condition': style_condition}

        # Loop through cells in the first (header) row and apply style_header(cell)
        for cell in ws[1]:  # First row for headers
            style_header(cell)

        # Loop through columns and apply formatting
        for i, label in enumerate(df_results.columns):
            # Set column width
            width = ProjectController.TABLE_COLUMNS.get(label, {}).get('width', None)
            if width:
                ws.column_dimensions[get_column_letter(i + 1)].width = width

            # Apply styles to the rest of the column
            for row in ws.iter_rows(min_row=2, min_col=i + 1, max_col=i + 1):
                for cell in row:
                    alignment = ProjectController.TABLE_COLUMNS.get(label, {}).get('alignment', None)
                    style = ProjectController.TABLE_COLUMNS.get(label, {}).get('style', None)

                    if alignment == 'l':
                        cell.alignment = Alignment(horizontal='left', vertical='bottom')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='bottom')

                    if style:
                        style_functions[style](cell)
                    style_general(cell)

        ws.auto_filter.ref = "A1:" + get_column_letter(ws.max_column) + "1"

        # Save the workbook
        wb.save(file_path)

    def get_governing_by_group(self):
        self.governing_items = {}
        if self.excel_tables.project_info['report_by_group'] == 'Worst-Case':
            # Create dictionary of {group: list of items in group}
            df = self.df_equipment.set_index(self.df_equipment.columns[0])

            self.group_dict = {group: list(item.index) for group, item in df.groupby('group')}
            self.group_dict['ungrouped'] = df[df['group'].isna()].index.tolist()

            # Create dictionary whose keys are governing items, and values are their corresponding group
            for group, eq_list in self.group_dict.items():

                if group == 'ungrouped':
                    for id in eq_list:
                        self.governing_items[id] = (None, 0)
                    continue

                # todo: [Input Validation]: check anchorage type (asure all equipment items are the same, maybe require same geometry)
                if self.df_equipment[self.df_equipment['equipment_id'] == eq_list[0]]['installation_type'].iloc[
                    0] == 'Base Anchored':
                    anchor_tensions = [self.items_for_report[eq_id].base_anchors.DCR for eq_id in eq_list]
                    anchor_tensions = [np.inf if val is None else val for val in
                                       anchor_tensions]  # if Tu_max is None, it is assumed that there was a failed spacing requirment, and this should be taken as the governing unit.
                    max_idx = np.argmax(anchor_tensions)
                    self.governing_items[eq_list[max_idx]] = (group, max_idx)
                #todo: Fix logic for wall wood fasteners
                elif self.df_equipment[self.df_equipment['equipment_id'] == eq_list[0]]['installation_type'].iloc[
                    0] in ['Wall Brackets', 'Wall Mounted']:
                    anchor_tensions = [
                        max([anchors.DCR for wall, anchors in self.items_for_report[eq_id].wall_anchors.items() if
                             anchors is not None] +
                            [backing.anchors_obj.DCR for backing in self.items_for_report[eq_id].wall_backing if
                             backing.anchors_obj is not None]) for eq_id in eq_list]
                    anchor_tensions = [np.inf if val is None else val for val in anchor_tensions]
                    max_idx = np.argmax(anchor_tensions)
                    self.governing_items[eq_list[max_idx]] = (group, max_idx)
                else:
                    raise Exception(
                        f'Installation type {self.df_equipment.loc[eq_list[0], "installation_type"]} for Equipment ID {eq_list[0]} not supported')
            self.governing_items = {
                k:self.governing_items[k]
                for k in self.items_for_report.keys()
                if k in self.governing_items}
        else:
            self.group_dict = None
            self.governing_items = {key: (None, 0) for key in self.items_for_report.keys()}

    def create_report(self):
        """ Creates a pdf report of EquipmentModel instances included in self.items_for_report"""

        report = EquipmentReport(self.excel_tables.project_info, self.items_for_report, self.governing_items,
                                 self.group_dict, pool=self.pool)
        # report.generate_pdf(self.output_dir)



