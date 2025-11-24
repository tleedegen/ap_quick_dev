import numpy as np

import pandas as pd


from dataclasses import fields

from PyQt5.uic.properties import Properties
from scipy.stats import false_discovery_control

from anchor_pro.model import ElementResults
from anchor_pro.project_controller.excel_importer import ExcelTablesImporter
from anchor_pro.reports.report import EquipmentReport
from anchor_pro.ap_types import FactorMethod
from anchor_pro.utilities import get_governing_result
import multiprocessing as mp
import itertools
import copy

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

import os


import anchor_pro.model as m
import anchor_pro.project_controller.factories as fact
from anchor_pro.project_controller.project_classes import *
import anchor_pro.elements.sms as sms


from typing import Optional, List, Tuple, Literal
from numpy.typing import NDArray

''' Dataclasses'''
Mode = Literal["Default", "Specified Product", "Product Group"]



'''Classes'''



class ProjectController:
    def __init__(self, excel_path, output_dir):
        # Inputs
        self.excel_path = excel_path
        self.output_dir = output_dir

        # Input Data
        self.excel_tables = None
        self.df_equipment = None  # Merge of equipment, base geometry, wall geometry, concrete tables
        self.get_model_inputs_from_excel_tables()

        # Models and Output
        # self.models: dict[str, m.EquipmentModel] = {}  # Dictionary of model objects (one for each row of the df_equipment)
        # self.analysis_runs_by_model: dict[str, list[AnalysisRun]] = {}
        # self.governing_analysis_run_by_model: dict[str, int] = {}  #{equipment id: index into analysis runs list}
        # self.groups: dict[str, list[str]] = {}  # {group name: list of equipment_ids
        # self.governing_model_by_group: dict[str, int] = {}  # {group name: index into list of equipment ids for}

        self.models: dict[ModelId, ModelRecord] = {}

    def get_model_inputs_from_excel_tables(self):
        """Performs appropriate lookups across dataframes (equipment table, base geometry, etc.) and assembles
        a dictionary of data series representing the required inputs for each equipment model.

        the dictionary contains the following keys:
        'equipment_data'
        'concrete_data'
        'base_geometry'
        'anchor_geometry'

        """
        # Read Excel File and Import Data Tables as DataFrames
        self.excel_tables = ExcelTablesImporter(self.excel_path)

        # Merge Tables
        df_equipment = self.excel_tables.df_equipment
        df_equipment = df_equipment[df_equipment['omit_calc'] != True]

        # Merge Base Geometry
        df_equipment = df_equipment.merge(self.excel_tables.df_base_geometry,
                                          left_on='base_geometry', right_on='Pattern Name',
                                          how='left')
        df_equipment = df_equipment.drop(columns=['Pattern Name'])

        # Merge Wall Geometry
        df_equipment = df_equipment.merge(self.excel_tables.df_wall_geometry,
                                          left_on='wall_geometry', right_on='Pattern Name',
                                          how='left', suffixes=('_base', '_wall'))
        df_equipment = df_equipment.drop(columns=['Pattern Name'])

        # Merge Base Material Data
        # todo: [New Base Materials] merge additional tables for other material types
        df_equipment = df_equipment.merge(self.excel_tables.df_concrete,
                                          left_on='base_mat_id', right_on='slab_id',
                                          how='left', suffixes=('_eq', '_base'))
        df_equipment = df_equipment.merge(self.excel_tables.df_wood,
                                          left_on='base_mat_id', right_on='wood_id',
                                          how='left', suffixes=('_eq', '_base'))

        # Merge Wall Material Data
        df_equipment = df_equipment.merge(self.excel_tables.df_walls,
                                          left_on='wall_id', right_on='wall_id',
                                          how='left', suffixes=('_base', '_wall'))
        df_equipment = df_equipment.merge(self.excel_tables.df_concrete,
                                          left_on='slab_id_wall', right_on='slab_id',
                                          how='left', suffixes=('_base', '_wall'))

        df_equipment['stud_or_blocking_key'] = (
            df_equipment['blocking_id'].astype(str).fillna(df_equipment['stud_id'].astype(str))
        )
        df_equipment = df_equipment.merge(
            self.excel_tables.df_wood,
            left_on='stud_or_blocking_key',right_on='wood_id',
            how='left', suffixes=('_base','_wall')).drop(columns=['stud_or_blocking_key'])

        ''' The decision is made deliberately not to merge anchor data or bracket data.
        These tables are independently quite large, and since checking by group may be selected,
        it seems preferable not to explicitly include these tables in df_equipment.'''

        self.df_equipment = df_equipment



    def run(self):

        # 1) Loop over equipment schedule
        for eq_idx, equipment_row in self.df_equipment.iterrows():
            # equipment_data = self.df_equipment.iloc[0]

            # 2) Create model object
            model_name = equipment_row['equipment_id']
            group_name = equipment_row['group']

            print(f'Creating Model: {model_name}')
            model, omit_bracket_output =fact.model_factory(equipment_row, self.excel_tables)\

            self.models[model_name] = ModelRecord(
                model=model,
                group = group_name,
                omit_bracket_output=omit_bracket_output)


            # 3) Loop over Hardware Selections
            plan = get_hardware_selection_plan(self.excel_tables, equipment_row, model)

            for base_anchor_id, bracket_id, wall_anchor_id, cxn_anchor_id in itertools.product(
                    plan.base_anchor_list, plan.bracket_list, plan.wall_anchor_list, plan.cxn_anchor_list
            ):
                print(f'Analyzing {model_name} with '
                      f'base anchor: {base_anchor_id}, '
                      f'wall bracket: {bracket_id}, '
                      f'wall anchor: {wall_anchor_id}, '
                      f'hardware fastener: {cxn_anchor_id}')

                selection = HardwareSelection(
                    base_anchor_id= base_anchor_id,
                    bracket_id=bracket_id,
                    wall_anchor_id=wall_anchor_id,
                    cxn_anchor_id=cxn_anchor_id)

                # 4) Set Hardware Data
                # 4a) Base Plate and Anchor Properties
                base_anchor_props = []
                base_plate_stiffness = []
                if base_anchor_id and model.install.base_material == m.BaseMaterial.concrete:
                    anchor_catalog = self.excel_tables.df_anchors
                    base_anchor_data = anchor_catalog[anchor_catalog['anchor_id'] == base_anchor_id].iloc[0]
                    for obj in model.elements.base_anchors:
                        prop, bp_stiffness, position_ok = fact.mechanical_anchor_props_factory(base_anchor_data, obj, return_bp_stiffness=True)
                        base_anchor_props.append(prop)
                        base_plate_stiffness.append(bp_stiffness)
                        # todo: figure out where ok goes (installation position)
                if base_anchor_id and model.install.base_material == m.BaseMaterial.wood:
                    #todo [Base Wood Anchors: hardware selection]
                    raise NotImplementedError

                # 4b) Wall Anchor Props and Bracket Props
                wall_anchor_props = []
                if wall_anchor_id and model.install.base_material == m.WallMaterial.concrete:
                    anchor_catalog = self.excel_tables.df_anchors
                    wall_anchor_data = anchor_catalog[anchor_catalog['anchor_id'] == wall_anchor_id].iloc[0]
                    for obj in model.elements.wall_anchors:
                        prop, position_ok = fact.mechanical_anchor_props_factory(wall_anchor_data, obj)
                        wall_anchor_props.append(prop)
                        # todo: figure out where ok goes (installation position)
                if wall_anchor_id and model.install.base_material == m.WallMaterial.wood:
                    # todo [Wall Wood Anchors: hardware selection]
                    raise NotImplementedError
                if wall_anchor_id and model.install.base_material == m.WallMaterial.cfs:
                    # todo [Wall SMS Anchors: hardware selection]
                    raise NotImplementedError

                # Wall Bracket Properties
                wall_bracket_props = []
                if bracket_id:
                    bracket_catalog = self.excel_tables.df_brackets_catalog
                    bracket_data = bracket_catalog[bracket_catalog['bracket_id']==bracket_id].iloc[0]
                    for bracket in model.elements.wall_brackets:
                        prop = fact.wall_bracket_props_factory(bracket_id,bracket_data,bracket.geo_props.wall_flexibility)
                        wall_bracket_props.append(prop)

                # Set All Properties
                model.set_model_data(
                    base_anchor_props=base_anchor_props,
                    base_plate_stiffness=base_plate_stiffness,
                    wall_anchor_props=wall_anchor_props,
                    wall_bracket_props=wall_bracket_props,
                    cxn_sms_size=cxn_anchor_id,
                    sms_catalog=self.excel_tables.sms_catalog)

                # 5) Perform Pre-Analysis Checks
                # todo: incorporate model_spacing check
                model.check_model_stability()
                if model.analysis_vars.omit_analysis:
                    self.models[model_name].analysis_runs.append(
                        AnalysisRun(
                            equipment_id=model.equipment_info.equipment_id,
                            hardware_selection=selection,
                            omit_analysis=True))
                    continue
                # 6) Run Analysis
                # Initialize solutions dicitionary
                solutions = {}

                # Extract previous cached solution (if applicable) as intial guess
                ''' If previous analysis runs exists, extract the run whose hardware has the closets stiffness 
                for use as the initial guess solution with the current hardware selection. 
                Code prioritizes base anchor stiffness, then wall bracket stiffness for comparission.'''
                if model.elements.base_anchors:
                    K_current = model.elements.base_anchors[0].anchor_props.K
                    stiffness_differences = [abs(run.results.base_anchors[0].K - K_current)
                                             for run in self.models[model_name].analysis_runs]
                elif model.elements.wall_brackets:
                    K_current = model.elements.wall_brackets[0].bracket_props.kn
                    stiffness_differences = [abs(run.results.wall_brackets[0].kn - K_current)
                                             for run in self.models[model_name].analysis_runs]
                else:
                    raise NotImplementedError("No stiffness comparison provided for model without either base anchors or wall brackets")

                for method in model.analysis_vars.factor_methods:
                    Fh, Fv = model.factored_loads.get(method)
                    previous_solutions = [run.solutions[method] for run in self.models[model_name].analysis_runs]

                    if previous_solutions:
                        idx_closest_K = min(range(len(stiffness_differences)), key=stiffness_differences.__getitem__)
                        trial_solution = previous_solutions[idx_closest_K]
                    else:
                        trial_solution=None

                    # Run analysis
                    solutions[method] = (model.analyze(Fh, Fv, method, initial_solution_cache=trial_solution))

                # 7) Evaluate Elements
                element_results = model.evaluate(solutions)

                self.models[model_name].analysis_runs.append(
                    AnalysisRun(
                        equipment_id=model.equipment_info.equipment_id,
                        hardware_selection=selection,
                        omit_analysis=False,
                        results=element_results,
                        solutions=solutions))

            # 9) Identify Governing and Optimum Runs
            self.models[model_name].governing_run = get_optimum_hardware(self.models[model_name])


    def results_to_dataframe(self) -> pd.DataFrame:
        """
        Assemble the summary DataFrame consumed by create_excel_summary_table(...)
        One row per AnalysisRun in self.analysis_runs.
        Leaves 'optimum' columns untouched (to be post-processed elsewhere).
        """

        cols = list(TABLE_COLUMNS.keys())

        def _empty_row():
            row = {}
            for c in cols:
                style = TABLE_COLUMNS[c].get("style")
                # default by style / name family
                if style in ("dcr",):
                    row[c] = np.nan
                elif style in ("result", "condition", "optimum"):
                    row[c] = None
                elif "Max" in c or c.endswith("(lbs)") or TABLE_COLUMNS[c].get("units"):
                    row[c] = np.nan
                else:
                    row[c] = ""
            return row

        def _nanmax_safe(arr):
            try:
                return float(np.nanmax(arr)) if arr is not None else np.nan
            except Exception:
                return np.nan

        rows = []
        for id, mrec in self.models.items():
            for run_idx, run in enumerate(mrec.analysis_runs):
                row = _empty_row()

                # ---------- General info ----------
                # Item / Group / Wp / Fp
                eq_id = getattr(run, "equipment_id", "")
                row["Item"] = eq_id

                mdl = mrec.model
                eqinfo = getattr(mdl, "equipment_info", None) if mdl else None
                eprops = getattr(mdl, "equipment_props", None) if mdl else None
                row["Group"] = getattr(eqinfo, "group", "-")
                row["Wp"] = getattr(eprops, "Wp", np.nan)

                # Fp: if your model or solutions expose it, fill it; otherwise leave NaN
                # Try LRFD first, else any solution with an attribute/field.
                row["Fp"] = mdl.fp_calc.Fp

                # ---------- Hardware selections ----------
                hs = getattr(run, "hardware_selection", {}) or {}
                row["Base Anchor"] = hs.base_anchor_id
                row["Hardware SMS"] = hs.cxn_anchor_id
                # row["Base Strap"] = hs.base_strap
                # row["Wall Bracket"] = hs.bracket_id
                #
                # row["Wall SMS"] = hs.wall_anchor_id
                # row["Wall Fastener"] = hs.get("wall_fastener_id", "")
                # row["Wall Anchor"] = hs.get("wall_anchor_id", "")

                # ---------- Results extraction ----------
                res = getattr(run, "results", None)

                # Base Anchors (Concrete)
                if res.base_anchors and mrec.model.install.base_material==m.BaseMaterial.concrete:
                    #Extract governing items
                    ba = max(res.base_anchors, key=lambda x: x.unity)  # Extract governing base anchor object
                    tg_idx = ba.governing_tension_group # governing tension group index
                    sg_idx = ba.governing_shear_group  # governing shear group index
                    anchor_idx = ba.governing_anchor_idx
                    theta_idx = ba.governing_theta_idx

                    tg = ba.tension_groups[tg_idx]  # Extract governing tension group
                    anchor_in_tg = tg.anchor_indices.index(anchor_idx)

                    row["Base Anchor Max Tension"] = ba.governing_tension
                    row["Base Thickness Check"] = ba.spacing_requirements.slab_thickness_ok
                    row["Base Spacing and Edge Checks"] = ba.spacing_requirements.slab_thickness_ok

                    for limit, calc in {
                        "Base Steel Tensile Strength": ba.steel_tension_calcs[tg_idx],
                        "Base Concrete Tension Breakout": ba.tension_breakout_calcs[tg_idx],
                        "Base Anchor Pullout": ba.anchor_pullout_calcs[tg_idx],
                        "Base Side Face Blowout": ba.side_face_blowout_calcs[tg_idx],
                        "Base Bond Strength": ba.bond_strength_calcs[tg_idx],
                        "Base Steel Shear Strength": ba.steel_shear_calcs[tg_idx],
                        "Base Shear Pryout": ba.shear_pryout_calcs[tg_idx]
                    }.items():
                        unities = getattr(calc, "unities", None)
                        if unities is not None:
                            unity = unities[anchor_in_tg, theta_idx]
                        else:
                            unity = "NA"
                        row[limit] = unity

                    if sg_idx > -1:  # Extract governing shear group
                        sg = ba.shear_groups[sg_idx]
                        anchor_in_sg = sg.anchor_indices.index(anchor_idx)
                        unities = getattr(ba.shear_breakout_calcs[sg_idx],"unities",None)
                        if unities is not None:
                            unity = unities[anchor_in_sg, theta_idx]
                            row["Base Shear Breakout"] = unity
                    else:
                        row["Base Shear Breakout"] = "NA"

                    row["Base Tension-Shear Interaction"] = ba.unity
                    row["Base Anchor OK"] = ba.unity <= 1.0
                    row['Optimum Base Anchor'] = (run_idx == mrec.governing_run) and (ba.unity <= 1.0)

                # Hardware SMS
                if res and getattr(res, "base_plate_fasteners", None):
                    sms0 = res.base_plate_fasteners[0]
                    # Max Tension / Shear
                    row["Hardware SMS Max Tension"] = _nanmax_safe(getattr(sms0, "tension", None))
                    row["Hardware SMS Max Shear"] = _nanmax_safe(getattr(sms0, "shear", None))
                    # DCRs
                    row["Hardware SMS Tension DCR"] = _nanmax_safe(getattr(sms0, "dcr_tension", None))
                    row["Hardware SMS Shear DCR"] = _nanmax_safe(getattr(sms0, "dcr_shear", None))
                    row["Hardware SMS OK"] = getattr(sms0, "ok", None)

                # Base Strap
                if res and getattr(res, "base_straps", None):
                    bs0 = res.base_straps[0]
                    row["Maximum Base Strap Tension"] = _nanmax_safe(getattr(bs0, "tension", None))
                    row["Base Strap DCR"] = _nanmax_safe(getattr(bs0, "dcr", None))
                    row["Base Strap OK"] = getattr(bs0, "ok", None)

                # Wall Bracket
                if res and getattr(res, "wall_brackets", None):
                    wb0 = res.wall_brackets[0]
                    row["Maximum Bracket Tension"] = _nanmax_safe(getattr(wb0, "tension", None))
                    row["Bracket DCR"] = _nanmax_safe(getattr(wb0, "dcr", None))
                    row["Bracket OK"] = getattr(wb0, "ok", None)

                # Wall SMS Anchors
                if res and getattr(res, "wall_sms", None):
                    wsm0 = res.wall_sms[0]
                    row["Wall SMS Max Tension"] = _nanmax_safe(getattr(wsm0, "tension", None))
                    row["Wall SMS Max Shear"] = _nanmax_safe(getattr(wsm0, "shear", None))
                    row["Wall SMS Tension DCR"] = _nanmax_safe(getattr(wsm0, "dcr_tension", None))
                    row["Wall SMS Shear DCR"] = _nanmax_safe(getattr(wsm0, "dcr_shear", None))
                    row["Wall SMS DCR"] = _nanmax_safe(getattr(wsm0, "dcr", None))
                    row["Wall SMS OK"] = getattr(wsm0, "ok", None)

                # Wall Wood Fastener
                if res and getattr(res, "wall_fasteners", None):
                    wf0 = res.wall_fasteners[0]
                    row["Wall Fastener Max Force"] = _nanmax_safe(getattr(wf0, "force", None))
                    row["Wall Fastener DCR"] = _nanmax_safe(getattr(wf0, "dcr", None))
                    row["Wall Fastener OK"] = getattr(wf0, "ok", None)

                # Wall Concrete Anchor
                if getattr(res, "wall_anchors", None) and (mrec.model.install.wall_material==m.WallMaterial.concrete):
                    # Extract governing items
                    wa = max(res.wall_anchors, key=lambda x: x.unity)  # Extract governing base anchor object
                    tg_idx = wa.governing_tension_group  # governing tension group index
                    sg_idx = wa.governing_shear_group  # governing shear group index
                    anchor_idx = wa.governing_anchor_idx
                    theta_idx = wa.governing_theta_idx

                    tg = wa.tension_groups[tg_idx]  # Extract governing tension group
                    anchor_in_tg = tg.anchor_indices.index(anchor_idx)

                    row["Wall Anchor Max Tension"] = wa.governing_tension
                    row["Wall Thickness Check"] = wa.spacing_requirements.slab_thickness_ok
                    row["Wall Spacing and Edge Checks"] = wa.spacing_requirements.slab_thickness_ok

                    for limit, calc in {
                        "Wall Steel Tensile Strength": wa.steel_tension_calcs[tg_idx],
                        "Wall Concrete Tension Breakout": wa.tension_breakout_calcs[tg_idx],
                        "Wall Anchor Pullout": wa.anchor_pullout_calcs[tg_idx],
                        "Wall Side Face Blowout": wa.side_face_blowout_calcs[tg_idx],
                        "Wall Bond Strength": wa.bond_strength_calcs[tg_idx],
                        "Wall Steel Shear Strength": wa.steel_shear_calcs[tg_idx],
                        "Wall Shear Pryout": wa.shear_pryout_calcs[tg_idx]
                    }.items():
                        unities = getattr(calc, "unities", None)
                        if unities is not None:
                            unity = unities[anchor_in_tg, theta_idx]
                        else:
                            unity = "NA"
                        row[limit] = unity

                    if sg_idx > -1:  # Extract governing shear group
                        sg = wa.shear_groups[sg_idx]
                        anchor_in_sg = sg.anchor_indices.index(anchor_idx)
                        unities = getattr(wa.shear_breakout_calcs[sg_idx], "unities", None)
                        if unities is not None:
                            unity = unities[anchor_in_sg, theta_idx]
                            row["Wall Shear Breakout"] = unity
                    else:
                        row["Wall Shear Breakout"] = "NA"

                    row["Wall Tension-Shear Interaction"] = wa.unity
                    row["Wall Anchor OK"] = wa.unity <= 1.0
                    row['Optimum Wall Anchor'] = (run_idx == mrec.governing_run) and (wa.unity <= 1.0)

                # Keep the row
                rows.append(row)

        return pd.DataFrame(rows, columns=cols)

    def create_excel_summary_table(self):
        df_results = self.results_to_dataframe()

        print('Creating summary table')
        # Export and Format Results Summary Table (using Openpyxl)
        file_path = os.path.join(self.output_dir, 'Results Summary New.xlsx')

        wb = Workbook()
        ws = wb.active

        # Populate the sheet with DataFrame excel_tables
        for r in dataframe_to_rows(df_results, index=False, header=True):
            ws.append(r)

        # Predefine Fill Colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FF7276", end_color="FF7276", fill_type="solid")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        def style_dcr(c):
            if isinstance(c.value, (int, float)) and c.value > 1:
                c.fill = red_fill

        def style_condition(c):
            if c.value == 'NG':
                c.fill = red_fill
            elif c.value == 'OK':
                c.fill = green_fill

        def style_result(c):
            if c.value:
                c.fill = green_fill
            else:
                c.fill = gray_fill

        def style_optimum(c):
            if c.value:
                c.fill = green_fill
            else:
                c.fill = gray_fill

        def style_header(c):
            c.fill = header_fill
            c.font = Font(color="FFFFFF")  # White Font
            c.border = thin_border
            c.alignment = Alignment(horizontal='left', vertical='bottom', text_rotation=45)

        def style_general(c):
            # Only apply number formatting if it's a number, not a boolean
            if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                c.number_format = '0.00'
            c.border = thin_border
            if c.value == 'NA' or c.value == "":
                c.fill = gray_fill

        style_functions = {'dcr': style_dcr,
                           'optimum': style_optimum,
                           'result': style_result,
                           'condition': style_condition}

        # Loop through cells in the first (header) row and apply style_header(cell)
        for cell in ws[1]:  # First row for headers
            style_header(cell)

        # Loop through columns and apply formatting
        for i, label in enumerate(df_results.columns):
            # Set column width
            width = TABLE_COLUMNS.get(label, {}).get('width', None)
            if width:
                ws.column_dimensions[get_column_letter(i + 1)].width = width

            # Apply styles to the rest of the column
            for row in ws.iter_rows(min_row=2, min_col=i + 1, max_col=i + 1):
                for cell in row:
                    alignment = TABLE_COLUMNS.get(label, {}).get('alignment', None)
                    style = TABLE_COLUMNS.get(label, {}).get('style', None)

                    if alignment == 'l':
                        cell.alignment = Alignment(horizontal='left', vertical='bottom')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='bottom')

                    if style:
                        style_functions[style](cell)
                    style_general(cell)

        ws.auto_filter.ref = "A1:" + get_column_letter(ws.max_column) + "1"

        # Save the workbook
        wb.save(file_path)

    def create_report(self):
        """ Creates a pdf report of EquipmentModel instances included in self.items_for_report"""

        # Set up parallel processing
        self.pool = None #todo fix for parallel processing. Review old code

        # Post-process Governing by group, etc.
        for id, mrec in self.models.items():
            # Set Props for stand-alone section
            if mrec.group is None:
                mrec.for_report = True
                mrec.report_section_name = f'{id} [{mrec.model.equipment_info.equipment_type}]'
            # Set Props for group section
            else:
                governing_group_items = get_governing_group_items(self.models)
                for group_name, model_name in governing_group_items.items():
                    self.models[model_name].for_report = True
                    self.models[model_name].report_section_name = group_name


        EquipmentReport(self.excel_tables.project_info, self.models, pool=self.pool)


'''Global Variables'''
TABLE_COLUMNS = {
        # GENERAL INFO
        'Item': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Group': {'units': None, 'width': 12, 'alignment': 'l', 'style': None},
        'Wp': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Fp': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        # BASE CONCRETE ANCHOR RESULTS
        'Base Anchor': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Base Anchor Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Base Thickness Check': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Base Spacing and Edge Checks': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Base Steel Tensile Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Concrete Tension Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Anchor Pullout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Side Face Blowout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Bond Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Steel Shear Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Shear Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Shear Pryout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Tension-Shear Interaction': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Anchor OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Base Anchor': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # HARDWARE FASTENERS
        'Hardware SMS': {'units': None, 'width': None, 'alignment': 'l', 'style': None},
        'Hardware SMS Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Hardware SMS Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Hardware SMS Tension DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Hardware SMS Shear DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Hardware SMS OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Hardware SMS': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # BASE STRAPS
        'Base Strap': {'units': None, 'width': 16, 'alignment': 'l', 'style': None},
        'Maximum Base Strap Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Base Strap DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Base Strap OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL BRACKETS
        'Wall Bracket': {'units': None, 'width': 16, 'alignment': 'l', 'style': None},
        'Maximum Bracket Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Bracket DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Bracket OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL SMS ANCHORS
        'Wall SMS': {'units': None, 'width': 8, 'alignment': 'l', 'style': None},
        'Wall SMS Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall SMS Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall SMS Tension DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS Shear DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall SMS OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Wall SMS': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'},
        # WALL WOOD FASTENER
        'Wall Fastener': {'units': None, 'width': 8, 'alignment': 'l', 'style': None},
        'Wall Fastener Max Force': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        #'Wall Fastner Max Shear': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall Fastener DCR': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Fastener OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        # WALL CONCRETE ANCHOR
        'Wall Anchor': {'units': None, 'width': 24, 'alignment': 'l', 'style': None},
        'Wall Anchor Max Tension': {'units': '(lbs)', 'width': None, 'alignment': None, 'style': None},
        'Wall Thickness Check': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Wall Spacing and Edge Checks': {'units': None, 'width': None, 'alignment': None, 'style': 'condition'},
        'Wall Steel Tensile Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Concrete Tension Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Anchor Pullout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Side Face Blowout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Bond Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Steel Shear Strength': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Shear Breakout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Shear Pryout': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Tension-Shear Interaction': {'units': None, 'width': None, 'alignment': None, 'style': 'dcr'},
        'Wall Anchor OK': {'units': None, 'width': None, 'alignment': None, 'style': 'result'},
        'Optimum Wall Anchor': {'units': None, 'width': None, 'alignment': None, 'style': 'optimum'}
    }





def get_hardware_selection_plan(
    excel_tables: ExcelTablesImporter,
    equipment_row: pd.Series,
    model: m.EquipmentModel,
) -> HardwareSelectionPlan:

    pi = excel_tables.project_info  # dict-like (modes live here)
    # df_anchors: pd.DataFrame = excel_tables.df_anchors
    # df_wall_brackets: pd.DataFrame = excel_tables.df_wall_brackets
    # df_wood_fasteners: pd.DataFrame = getattr(excel_tables, "df_wood_fasteners", pd.DataFrame())
    # bracket_groups_list: list[str] = excel_tables.bracket_groups_list
    # df_anchor_groups: pd.DataFrame = getattr(excel_tables, "df_product_groups", pd.DataFrame())  # optional

    # Read user-provided specific products / groups off the row (if present)
    base_anchor_id      = equipment_row.get("base_anchor_id")
    base_anchor_group   = equipment_row.get("base_anchor_group")
    bracket_id          = equipment_row.get("bracket_id")
    bracket_group       = equipment_row.get("bracket_group")
    wall_anchor_id      = equipment_row.get("wall_anchor_id")
    wall_anchor_group   = equipment_row.get("wall_anchor_group")
    wall_fastener_id    = equipment_row.get("wall_fastener_id")
    wall_fastener_group = equipment_row.get("wall_fastener_group")
    wall_sms_id         = equipment_row.get("wall_sms_id")
    wall_sms_group      = equipment_row.get("wall_sms_group")
    cxn_sms_id = equipment_row.get("cxn_sms_id")


    # Modes from project_info (fall back to "Default")
    base_concrete_mode   = pi.get("base_concrete_product_mode", "Default")
    hardware_sms_mode    = pi.get("hardware_sms_product_mode", "Default")
    wall_bracket_mode    = pi.get("wall_bracket_product_mode", "Default")
    wall_concrete_mode   = pi.get("wall_concrete_product_mode", "Default")
    wall_sms_mode        = pi.get("wall_sms_product_mode", "Default")
    wall_wood_mode       = pi.get("wall_wood_product_mode", "Default")

    has_base_anchors = model.elements.base_anchors
    has_wall_brackets = model.elements.wall_brackets
    has_wall_anchors = model.elements.wall_anchors
    has_sms_cxn = model.elements.base_plate_fasteners or model.elements.wall_bracket_fasteners

    # 1) Base anchors (concrete only in this example)
    if has_base_anchors and model.install.base_material == m.BaseMaterial.concrete:
        base_anchor_list = _resolve_concrete_anchor_list(
            mode=base_concrete_mode,
            specified_id=base_anchor_id,
            group_name=base_anchor_group,
            member_type=model.elements.base_anchors[0].concrete_props.profile,              # "Slab" | "Filled Deck"
            df_anchors=excel_tables.df_anchors,
            df_anchor_groups=excel_tables.df_product_groups
        )
    else:
        base_anchor_list = [None]

    # 2) Brackets
    if has_wall_brackets:
        bracket_list = _resolve_bracket_list(
            mode=wall_bracket_mode,
            specified_id=bracket_id,
            group_name=bracket_group,
            df_wall_brackets=excel_tables.df_brackets_catalog,
            bracket_groups_list=excel_tables.bracket_group_list,
        )
    else:
        bracket_list = [None]

    # 3) Wall anchors (depends on wall type)
    if has_wall_anchors:
        if model.install.wall_material in (m.WallMaterial.concrete, m.WallMaterial.cmu):
            wall_anchor_list = _resolve_concrete_anchor_list(
                mode=wall_concrete_mode,
                specified_id=wall_anchor_id,
                group_name=wall_anchor_group,
                member_type=model.elements.wall_anchors[0].concrete_props.profile,  # "Slab" | "Filled Deck"
                df_anchors=excel_tables.df_anchors,
                df_anchor_groups=excel_tables.df_product_groups
            )
    #     elif ms.wall_type == "Metal Stud":
    #         wall_anchor_list = _resolve_wall_sms_list(
    #             mode=wall_sms_mode,
    #             specified_id=wall_sms_id,
    #             group_name=wall_sms_group,
    #         )
    #     elif ms.wall_type == "Wood Stud":
    #         wall_anchor_list = _resolve_wall_wood_fastener_list(
    #             mode=wall_wood_mode,
    #             specified_id=wall_fastener_id,
    #             group_name=wall_fastener_group,
    #             df_wood_fasteners=df_wood_fasteners,
    #         )
    #     else:
    #         wall_anchor_list = [None]
    else:
        wall_anchor_list = [None]


    # 4) Connection SMS list (for hardware connection node), if present
    # cxn_anchor_list = _resolve_connection_sms_list(hardware_sms_mode) if ms.has_connections else [None]
    if has_sms_cxn:
        cxn_anchor_list = _resolve_sms_list(hardware_sms_mode,cxn_sms_id)
    else:
        cxn_anchor_list = [None]

    # Final plan
    return HardwareSelectionPlan(
        base_anchor_list=base_anchor_list or [None],
        bracket_list=bracket_list or [None],
        wall_anchor_list=wall_anchor_list or [None],
        cxn_anchor_list=cxn_anchor_list or [None],
    )

def _resolve_concrete_anchor_list(
        mode: str,
        specified_id: Optional[str],
        group_name: Optional[str],
        member_type: str,
        df_anchors: pd.DataFrame,
        df_anchor_groups: pd.DataFrame,
    ) -> list[str] | list[None]:

    if mode in ("Default", "Specified Product") and specified_id:
        return [specified_id]
    if mode in ("Default", "Product Group") and group_name:
        product_col = df_anchor_groups.columns[0]
        mask = df_anchor_groups[group_name].fillna(False).astype(bool)
        anchor_products = df_anchor_groups.loc[mask, product_col]
        filtered_anchors_df = df_anchors[df_anchors['product'].isin(anchor_products)]
        if member_type == 'Slab':
            filtered_anchors_df = filtered_anchors_df[filtered_anchors_df['slab_ok']]
        elif member_type == 'Filled Deck':
            filtered_anchors_df = filtered_anchors_df[filtered_anchors_df['deck_top_ok']]
        return filtered_anchors_df['anchor_id'].tolist()
    raise Exception('Must specify anchor product group or product id.')

def _resolve_bracket_list(
        mode: str,
        specified_id: Optional[str],
        group_name: Optional[str],
        df_wall_brackets: pd.DataFrame,
        bracket_groups_list: list[str],
    ) -> list[str] | list[None]:
    if mode in ("Default", "Specified Product") and specified_id:
        return [specified_id]
    if mode in ("Default", "Product Group") and group_name:
        idx = bracket_groups_list.index(group_name) + 1
        col = f"group_{idx}"
        if col in df_wall_brackets.columns:
            filtered_bracket_df = df_wall_brackets[df_wall_brackets[col] == True]
            return filtered_bracket_df['bracket_id'].tolist()
    return [None]

def _resolve_sms_list(
        mode: str,
        specified_id: Optional[str]=None
    ):

    sms_list = ['No. 14', 'No. 12', 'No. 10', 'No. 8', 'No. 6']


    if mode == "Specified Product":
        if not specified_id in sms_list:
            specified_id = 'No. 12'
        return [specified_id]
    if mode == "Product Group":
        return sms_list
    if mode == "Default":
        if not specified_id in sms_list:
            return sms_list

    return [specified_id]

def get_optimum_hardware(
        model_record:ModelRecord
    ):

    model = model_record.model
    if model.elements.base_anchors and model.install.base_material==m.BaseMaterial.concrete:
        governing_run = _optimum_concrete_anchor(model_record,'base_anchors')
    elif model.elements.wall_anchors and model.install.wall_material==m.WallMaterial.concrete:
        governing_run = _optimum_concrete_anchor(model_record, 'wall_anchors')
    else:
        raise NotImplementedError("No method defined for 'optimum' hardware selection")

    return governing_run



    base_anchor_max_unities = [get_governing_result(run.results.base_anchors)[0].unity for run in model_record.analysis_runs]
    cost_rank_and_hardware = [(run.hardware_selection.base_anchor_id,
                               get_governing_result(run.results.base_anchors)[0].unity) for run in model_record.analysis_runs if get_governing_result(run.results.base_anchors)[0].unity<=1.0]




def get_governing_group_items(model_records: dict):
    # Collect list of group names
    group_names = set([mrec.group for name, mrec in model_records.items()])

    # Loop Over Groups and identify governing item
    group_governing_item = {}
    for group in group_names:
        group_governing_runs = [(name, mrec.analysis_runs[mrec.governing_run].results) for name, mrec in model_records.items() if mrec.group == group]
        group_governing_item[group] = max(group_governing_runs, key=lambda x: x[1].max_unity)[0]

    return group_governing_item

def _optimum_concrete_anchor(model_record: ModelRecord, element_type: Literal['base_anchors','wall_anchors']):
    ranked_ok_runs = []  # Successful Runs
    ranked_overstressed_runs = []  # Runs in which base anchors are overstressed
    ranked_spacing_ng_runs = []  # Runs in which base anchor spacing is no good

    run_indices = []
    for idx_run, run in enumerate(model_record.analysis_runs):
        ba, ba_idx = get_governing_result(getattr(run.results, element_type))
        if not ba.spacing_requirements.ok:
            ranked_spacing_ng_runs.append((idx_run, ba.cost_rank))
        elif ba.unity > 1.0:
            ranked_overstressed_runs.append((idx_run, ba.unity))
        else:
            ranked_ok_runs.append((idx_run, ba.cost_rank))

    if ranked_ok_runs:
        # Case 1, successful anchor, take governing as lowest-cost anchor
        governing_run = min(ranked_ok_runs, key=lambda idx_cost: idx_cost[1])[0]
    elif ranked_overstressed_runs:
        #  Case 2, Anchors meet spacing requirements, but overstressed. Take governing as lowest unity value.
        governing_run = min(ranked_overstressed_runs, key=lambda idx_unity: idx_unity[1])[0]
    else:
        # Case 3: No anchors meet spacing requirements, take lowest-cost anchor
        governing_run = min(ranked_spacing_ng_runs, key=lambda idx_cost: idx_cost[1])[0]

    return governing_run